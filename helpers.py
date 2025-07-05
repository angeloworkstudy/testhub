# helpers.py
from database import get_connection

def ajouter_employe(nom, prenom, fonction, departement, type_contrat, date_enregistrement):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO EMPLOYES (nom, prenom, fonction, departement, type_contrat, date_enregistrement)
        VALUES (?, ?, ?, ?, ?, ?)
    """, (nom, prenom, fonction, departement, type_contrat, date_enregistrement))
    conn.commit()
    conn.close()

def recuperer_employes():
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT * FROM EMPLOYES")
    resultats = cur.fetchall()
    conn.close()
    return resultats

def supprimer_employe(id_employe):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("DELETE FROM EMPLOYES WHERE id_employe = ?", (id_employe,))
    conn.commit()
    conn.close()

def modifier_employe(id_employe, nom, prenom, fonction, departement, type_contrat):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        UPDATE EMPLOYES SET nom=?, prenom=?, fonction=?, departement=?, type_contrat=?
        WHERE id_employe=?
    """, (nom, prenom, fonction, departement, type_contrat, id_employe))
    conn.commit()
    conn.close()


def get_fiche_employe(id_employe):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT E.nom, E.prenom, E.fonction, E.departement, E.type_contrat,
               F.titre, S.date_formation, SU.statut, SU.date_validation, FO.nom
        FROM EMPLOYES E
        JOIN SUIVI SU ON E.id_employe = SU.id_employe
        JOIN SESSIONS S ON SU.id_session = S.id_session
        JOIN FORMATIONS F ON S.id_formation = F.id_formation
        JOIN FORMATEURS FO ON S.id_formateur = FO.id_formateur
        WHERE E.id_employe = ?
        ORDER BY S.date_formation DESC
    """, (id_employe,))
    resultats = cur.fetchall()
    conn.close()
    return resultats



# 📚 FORMATIONS

def ajouter_formation(titre, duree, unite_duree, date_enregistrement):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO FORMATIONS (titre, duree, unite_duree, date_enregistrement)
        VALUES (?, ?, ?, ?)
    """, (titre, duree, unite_duree, date_enregistrement))
    conn.commit()
    conn.close()

def recuperer_formations():
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT * FROM FORMATIONS")
    resultats = cur.fetchall()
    conn.close()
    return resultats

def supprimer_formation(id_formation):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("DELETE FROM FORMATIONS WHERE id_formation = ?", (id_formation,))
    conn.commit()
    conn.close()

def modifier_formation(id_formation, titre, duree, unite_duree):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        UPDATE FORMATIONS SET titre=?, duree=?, unite_duree=?
        WHERE id_formation=?
    """, (titre, duree, unite_duree, id_formation))
    conn.commit()
    conn.close()


# 🧑‍🏫 FORMATEURS

def ajouter_formateur(nom, responsable, email, telephone, adresse, date_enregistrement):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO FORMATEURS (nom, responsable, email, telephone, adresse, date_enregistrement)
        VALUES (?, ?, ?, ?, ?, ?)
    """, (nom, responsable, email, telephone, adresse, date_enregistrement))
    conn.commit()
    conn.close()

def recuperer_formateurs():
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT * FROM FORMATEURS")
    resultats = cur.fetchall()
    conn.close()
    return resultats

def supprimer_formateur(id_formateur):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("DELETE FROM FORMATEURS WHERE id_formateur = ?", (id_formateur,))
    conn.commit()
    conn.close()

def modifier_formateur(id_formateur, nom, responsable, email, telephone, adresse):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        UPDATE FORMATEURS
        SET nom=?, responsable=?, email=?, telephone=?, adresse=?
        WHERE id_formateur=?
    """, (nom, responsable, email, telephone, adresse, id_formateur))
    conn.commit()
    conn.close()




# 🗓️ SESSIONS

def ajouter_session(id_formation, id_formateur, date_formation, prix, paiement_effectue):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO SESSIONS (id_formation, id_formateur, date_formation, prix, paiement_effectue)
        VALUES (?, ?, ?, ?, ?)
    """, (id_formation, id_formateur, date_formation, prix, paiement_effectue))
    conn.commit()
    conn.close()

def recuperer_sessions():
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT SESSIONS.id_session, FORMATIONS.titre, FORMATEURS.nom, SESSIONS.date_formation, SESSIONS.prix, SESSIONS.paiement_effectue
        FROM SESSIONS
        JOIN FORMATIONS ON SESSIONS.id_formation = FORMATIONS.id_formation
        JOIN FORMATEURS ON SESSIONS.id_formateur = FORMATEURS.id_formateur
        ORDER BY SESSIONS.date_formation DESC
    """)
    sessions = cur.fetchall()
    conn.close()
    return sessions

def supprimer_session(id_session):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("DELETE FROM SESSIONS WHERE id_session = ?", (id_session,))
    conn.commit()
    conn.close()


# ✅ SUIVI

def ajouter_suivi(id_employe, id_session, statut, date_validation, date_enregistrement):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO SUIVI (id_employe, id_session, statut, date_validation, date_enregistrement)
        VALUES (?, ?, ?, ?, ?)
    """, (id_employe, id_session, statut, date_validation, date_enregistrement))
    conn.commit()
    conn.close()

def recuperer_suivis():
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT SUIVI.id_suivi, EMPLOYES.nom, EMPLOYES.prenom, FORMATIONS.titre,
               FORMATEURS.nom, SESSIONS.date_formation, SUIVI.statut, SUIVI.date_validation
        FROM SUIVI
        JOIN EMPLOYES ON SUIVI.id_employe = EMPLOYES.id_employe
        JOIN SESSIONS ON SUIVI.id_session = SESSIONS.id_session
        JOIN FORMATIONS ON SESSIONS.id_formation = FORMATIONS.id_formation
        JOIN FORMATEURS ON SESSIONS.id_formateur = FORMATEURS.id_formateur
        ORDER BY SESSIONS.date_formation DESC
    """)
    resultats = cur.fetchall()
    conn.close()
    return resultats

def supprimer_suivi(id_suivi):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("DELETE FROM SUIVI WHERE id_suivi = ?", (id_suivi,))
    conn.commit()
    conn.close()











import pandas as pd
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook

def exporter_employes_excel(nom_fichier="employes_export.xlsx"):
    employes = recuperer_employes()
    colonnes = ["Nom", "Prénom", "Fonction", "Département", "Type de contrat"]
    df = pd.DataFrame(employes, columns=[
        "ID", "Nom", "Prénom", "Fonction", "Département", "Type de contrat", "Date enregistrement"
    ])[colonnes]

    wb = Workbook()
    ws = wb.active
    ws.title = "Employés"

    # ➕ Ajout des lignes depuis le DataFrame (sans index)
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)

            # 🎨 Style d'en-tête
            if r_idx == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4F81BD", fill_type="solid")
                thin = Side(border_style="thin", color="000000")
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            else:
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                     top=Side(style='thin'), bottom=Side(style='thin'))

    # 🔧 Ajuster largeur auto
    for column_cells in ws.columns:
        max_len = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = max_len + 2

    wb.save(nom_fichier)
    return nom_fichier


import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import os

def exporter_formations_excel(nom_fichier="data/export_formations.xlsx"):
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl import Workbook

    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT titre, duree, unite_duree FROM FORMATIONS")
    donnees = cur.fetchall()
    conn.close()

    # 🔄 Création DataFrame
    df = pd.DataFrame(donnees, columns=["Titre", "Durée", "Unité"])

    wb = Workbook()
    ws = wb.active
    ws.title = "Formations"

    # ➕ Ajout des lignes avec style
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)

            # En-tête
            if r_idx == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4F81BD", fill_type="solid")
                thin = Side(border_style="thin", color="000000")
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = Alignment(horizontal="center")
            else:
                # Bordures et alignement
                thin = Side(border_style="thin", color="000000")
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = Alignment(horizontal="center")

    # 🔧 Ajuster largeur auto
    for column_cells in ws.columns:
        max_len = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = max_len + 2

    wb.save(nom_fichier)
    return nom_fichier





def exporter_formateurs_excel(nom_fichier="data/formateurs_export.xlsx"):
    formateurs = recuperer_formateurs()
    colonnes = ["Nom", "Responsable", "Email", "Téléphone", "Adresse"]
    df = pd.DataFrame(formateurs, columns=[
        "ID", "Nom", "Responsable", "Email", "Téléphone", "Adresse", "Date enregistrement"
    ])[colonnes]

    wb = Workbook()
    ws = wb.active
    ws.title = "Formateurs"

    # ➕ Ajout des lignes depuis le DataFrame
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4F81BD", fill_type="solid")
                thin = Side(border_style="thin", color="000000")
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            else:
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                     top=Side(style='thin'), bottom=Side(style='thin'))

    # 🔧 Ajuster largeur des colonnes
    for column_cells in ws.columns:
        max_len = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = max_len + 2

    wb.save(nom_fichier)
    return nom_fichier




def recuperer_suivis_groupes_par_session():
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT 
            SESSIONS.id_session,
            FORMATIONS.titre,
            FORMATEURS.nom,
            SESSIONS.date_formation,
            EMPLOYES.nom,
            EMPLOYES.prenom,
            SUIVI.statut,
            SUIVI.date_validation
        FROM SUIVI
        JOIN EMPLOYES ON SUIVI.id_employe = EMPLOYES.id_employe
        JOIN SESSIONS ON SUIVI.id_session = SESSIONS.id_session
        JOIN FORMATIONS ON SESSIONS.id_formation = FORMATIONS.id_formation
        JOIN FORMATEURS ON SESSIONS.id_formateur = FORMATEURS.id_formateur
        ORDER BY SESSIONS.date_formation DESC
    """)
    rows = cur.fetchall()
    conn.close()

    groupes = {}
    for row in rows:
        id_session = row[0]
        if id_session not in groupes:
            groupes[id_session] = {
                "titre": row[1],
                "formateur": row[2],
                "date": row[3],
                "participants": []
            }
        groupes[id_session]["participants"].append(row[4:])
    return groupes



def exporter_suivis_excel(df, nom_fichier="data/suivis_filtrés.xlsx"):
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl import Workbook

    # 🧽 Supprimer la colonne "Validation" si elle existe
    if "Validation" in df.columns:
        df = df.drop(columns=["Validation"])

    wb = Workbook()
    ws = wb.active
    ws.title = "Suivis filtrés"

    # ➕ Ajouter les lignes avec style
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)

            # 🎨 En-tête
            if r_idx == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4F81BD", fill_type="solid")
                thin = Side(border_style="thin", color="000000")
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = Alignment(horizontal="center")
            else:
                # 📏 Bordures + centrage
                thin = Side(border_style="thin", color="000000")
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = Alignment(horizontal="center")

    # 📐 Ajuster largeur des colonnes
    for column_cells in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = max_len + 2

    wb.save(nom_fichier)
    return nom_fichier
